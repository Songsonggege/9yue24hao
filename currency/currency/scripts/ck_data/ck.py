# author:胡已源
# datetime:2019/9/3 下午2:48
# software: PyCharm
import hashlib
import time

import pymongo
import json
import requests
from openpyxl import load_workbook
import toolkit


def get_ck_client():
    client = pymongo.MongoClient(
        "mongodb://ck:kelrjnQKYI@172.16.40.2:27701/ck"
    )

    db = client['ck']
    collection = db['s_capital']
    return collection


def get_company_list(excel_path):
    print("开始解析" + str(excel_path) + "文件中的数据")
    data_list = []
    excel = load_workbook(excel_path)
    sheet = excel.get_sheet_by_name("科创板公司代码列表")

    rows_num = sheet.max_row
    for i in range(2, rows_num + 1):
        text = sheet.cell(row=i, column=1).value.strip()
        data_list.append(text)

    return data_list


def main():
    excel = "/home/dev/PycharmProjects/script/currency/currency/data/CK-主板-新三板-债券V1.2.xlsx"
    data_list = get_company_list(excel)
    print(data_list)

    collection = get_ck_client()

    # data_list = ['A15257', 'A19297', 'A19298']
    for code in data_list:
        print("当前获取的code:", code)
        url = "http://app2.jg.eastmoney.com/stock/F9/GetStockStructureInfo?securityCode={}&startDate=1900-01-01&endDate=2019-12-31&order=desc&unitType=1&clientId=1".format(code)

        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36',
        }

        response = requests.get(url, headers=headers)
        try:
            json_object = json.loads(json.loads(response.text))

            m = hashlib.md5()
            m.update(code.encode("utf-8"))
            data = {
                "_id": m.hexdigest(),
                "code": code,
                "crawler_user": "胡已圆",
                "crawl_time": time.time(),
                "crawl_format_time": time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())),
                "content": json_object,
                "source_url": url
            }

            mongo_update_instruction = {
                "filter": {"_id": data['_id']},
                "update": {
                    "$set": data.copy()
                },
                "upsert": True
            }

            collection.update_one(**mongo_update_instruction)
        except:
            print("当前code无数据:", code)
            # print(response.text)


if __name__ == '__main__':
    main()
