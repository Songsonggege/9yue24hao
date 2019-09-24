# author:胡已源
# datetime:2019/9/11 下午2:06
# software: PyCharm
"""
 ck 数据-主营构成
"""
import hashlib
import time
import pymongo
import json
import requests
from openpyxl import load_workbook
import toolkit

s = requests.session()


def get_ck_client(collection_name):
    client = pymongo.MongoClient(
        "mongodb://ck:kelrjnQKYI@172.16.40.2:27701/ck"
    )

    db = client['ck']
    collection = db[collection_name]
    return collection


def start_url():
    url = "http://app2.jg.eastmoney.com/stock/f9/MainCompositionByRegion?securityCode={}.SH&industryType=1&showBlank=false&order=desc&startDate=1900-01-01&endDate=2019-12-31&reportPeriodType=1%2C5%2C3%2C6&exchangeRate=1&exchangeRateLabel=CNY&unitType=1&clientId=1"

    url2 = "http://app2.jg.eastmoney.com/stock/f9/MainCompositionByRegion?securityCode={}.SH&industryType=2&showBlank=false&order=desc&startDate=1900-01-01&endDate=2019-12-31&reportPeriodType=1%2C5%2C3%2C6&exchangeRate=1&exchangeRateLabel=CNY&unitType=1&clientId=1"
    return url, url2


def get_company_list(excel_path='/home/dev/PycharmProjects/script/currency/currency/data/CK-主板-新三板-债券V1.2.xlsx'):
    print("开始解析" + str(excel_path) + "文件中的数据")
    data_list = []
    excel = load_workbook(excel_path)
    sheet = excel.get_sheet_by_name("科创板公司代码列表")

    rows_num = sheet.max_row
    for i in range(2, rows_num + 1):
        text = sheet.cell(row=i, column=1).value.strip()
        data_list.append(text)

    return data_list


def main():
    url, url2 = start_url()

    data_list = get_company_list()

    collection = get_ck_client('s_constitute_region')

    for code in data_list:
        new_url = url.format(code)
        new_url2 = url2.format(code)
        print(new_url)
        print("当前获取的code:", code)

        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36',
        }

        response = s.get(new_url, headers=headers)
        response2 = s.get(new_url2, headers=headers)

        content = []
        if response.text:
            json_object = json.loads(json.loads(response.text))
            content.append(json_object)

        if response2.text:
            json_object_2 = json.loads(json.loads(response2.text))
            content.append(json_object_2)

        if content:
            m = hashlib.md5()
            m.update(code.encode("utf-8"))
            data = {
                "_id": m.hexdigest(),
                "code": code,
                "crawler_user": "胡已圆",
                "crawl_time": time.time(),
                "crawl_format_time": time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())),
                "content": content,
                "source_url": new_url
            }

            print(json.dumps(data, ensure_ascii=False))

            mongo_update_instruction = {
                "filter": {"_id": data['_id']},
                "update": {
                    "$set": data.copy()
                },
                "upsert": True
            }

            collection.update_one(**mongo_update_instruction)


if __name__ == '__main__':
    main()
